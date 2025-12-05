import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from num2words import num2words
import io
import os
import tempfile
import pandas as pd

# Page configuration for mobile-friendly design
st.set_page_config(
    page_title="Generator Faktur",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS for professional modern mobile-friendly design with animations
st.markdown("""
    <style>
        /* Анимации */
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        @keyframes slideIn {
            from { opacity: 0; transform: translateX(-20px); }
            to { opacity: 1; transform: translateX(0); }
        }
        
        @keyframes pulse {
            0%, 100% { transform: scale(1); }
            50% { transform: scale(1.05); }
        }
        
        @keyframes glow {
            0%, 100% { box-shadow: 0 0 5px rgba(102, 126, 234, 0.5); }
            50% { box-shadow: 0 0 20px rgba(102, 126, 234, 0.8); }
        }
        
        /* Основные настройки */
        .main > div {
            padding-top: 1rem;
            max-width: 100%;
            animation: fadeIn 0.5s ease-in;
        }
        
        /* Крупные поля ввода для мобильных */
        .stNumberInput > div > div > input,
        .stTextInput > div > div > input,
        .stSelectbox > div > div > select {
            font-size: 18px !important;
            padding: 12px !important;
            min-height: 48px !important;
            border-radius: 8px !important;
            border: 2px solid #e0e0e0 !important;
            transition: all 0.3s ease !important;
            animation: slideIn 0.4s ease-out;
        }
        
        .stNumberInput > div > div > input:focus,
        .stTextInput > div > div > input:focus,
        .stSelectbox > div > div > select:focus {
            border-color: #667eea !important;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1) !important;
            outline: none !important;
        }
        
        /* Крупные кнопки с градиентами */
        .stButton > button {
            font-size: 18px !important;
            padding: 14px 24px !important;
            min-height: 52px !important;
            border-radius: 12px !important;
            font-weight: 600 !important;
            transition: all 0.3s ease !important;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1) !important;
            animation: fadeIn 0.6s ease-out;
        }
        
        .stButton > button:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 6px 12px rgba(0,0,0,0.15) !important;
        }
        
        .stButton > button[kind="primary"] {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
            border: none !important;
        }
        
        .stButton > button[kind="primary"]:hover {
            background: linear-gradient(135deg, #764ba2 0%, #667eea 100%) !important;
            animation: pulse 0.6s ease-in-out;
        }
        
        .stButton > button[kind="secondary"] {
            background: white !important;
            border: 2px solid #e0e0e0 !important;
            color: #333 !important;
        }
        
        .stButton > button[kind="secondary"]:hover {
            border-color: #667eea !important;
            color: #667eea !important;
            transform: translateY(-2px) !important;
        }
        
        /* Карточки для клиник */
        .clinic-card {
            padding: 20px;
            border-radius: 16px;
            border: 3px solid #e0e0e0;
            margin: 10px 0;
            cursor: pointer;
            transition: all 0.3s ease;
            background: white;
            animation: fadeIn 0.5s ease-out;
        }
        .clinic-card:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }
        .clinic-card.selected {
            border-color: #1f77b4;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            box-shadow: 0 4px 20px rgba(102, 126, 234, 0.4);
            animation: glow 2s ease-in-out infinite;
        }
        
        /* Секции с анимацией */
        .section {
            background: #f8f9fa;
            padding: 20px;
            border-radius: 12px;
            margin: 20px 0;
            animation: fadeIn 0.6s ease-out;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }
        
        /* Метрики с анимацией */
        [data-testid="stMetricValue"] {
            font-size: 28px !important;
            font-weight: 700 !important;
            transition: all 0.3s ease !important;
        }
        
        [data-testid="stMetricValue"]:hover {
            transform: scale(1.1);
        }
        
        /* Успешные сообщения */
        .stSuccess {
            animation: slideIn 0.4s ease-out;
            border-radius: 8px;
            padding: 12px;
        }
        
        /* Информационные блоки */
        .stInfo {
            animation: fadeIn 0.5s ease-out;
            border-radius: 8px;
        }
        
        /* Мобильная оптимизация */
        @media (max-width: 768px) {
            .main > div {
                padding-left: 0.5rem;
                padding-right: 0.5rem;
            }
            .stButton > button {
                width: 100% !important;
            }
            h1 {
                font-size: 24px !important;
            }
        }
        
        /* Скрыть лишнее */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        
        /* Плавные переходы для всех элементов */
        * {
            transition: background-color 0.3s ease, color 0.3s ease, border-color 0.3s ease;
        }
        
        /* Анимация для заголовков */
        h1, h2, h3 {
            animation: fadeIn 0.6s ease-out;
        }
        
        /* Улучшенные метрики */
        [data-testid="stMetricContainer"] {
            background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
            padding: 15px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            animation: fadeIn 0.7s ease-out;
            transition: transform 0.3s ease;
        }
        
        [data-testid="stMetricContainer"]:hover {
            transform: translateY(-3px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }
        
        /* Анимация для кнопок выбора клиники */
        .stButton > button[key*="clinic"] {
            animation: fadeIn 0.5s ease-out;
            position: relative;
            overflow: hidden;
        }
        
        .stButton > button[key*="clinic"]:hover::before {
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.3), transparent);
            animation: shine 0.5s ease-in-out;
        }
        
        @keyframes shine {
            to { left: 100%; }
        }
        
        /* Анимация для полей ввода при фокусе */
        .stTextInput > div > div > input:focus,
        .stNumberInput > div > div > input:focus {
            animation: pulse 0.3s ease-in-out;
        }
        
        /* Улучшенные сообщения */
        .stSuccess, .stInfo, .stWarning, .stError {
            animation: slideIn 0.4s ease-out;
            border-radius: 8px;
            padding: 12px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        
        /* Прогресс спиннера */
        .stSpinner > div {
            animation: pulse 1s ease-in-out infinite;
        }
    </style>
""", unsafe_allow_html=True)

# Hardcoded data - Seller
SELLER_NAME = 'Praktyka Lekarska „Salutem" Iryna Berehova'
SELLER_ADDRESS = "ul. Okręg Wieleński 4a/1, 64-410 Sieraków"
SELLER_NIP = "7882010121"
SELLER_REGON = "388783174"

BANK_ACCOUNT = "76124065531111001128223126"
PAYMENT_METHOD = "przelew"
PAYMENT_TERM = "wg umowy"

NOTES = "SPRZEDAWCA NIE JEST PŁATNIKIEM PODATKU VAT. Usługi zwolnione na podstawie art. 43 ust 1 pkt. 18 Ustawy o podatku od towaru i usług (VAT)."

# Клиники (Buyers)
CLINICS = {
    "miedzychod": {
        "name": "SAMODZIELNY PUBLICZNY ZAKŁAD OPIEKI ZDROWOTNEJ W MIĘDZYCHODZIE",
        "address_line1": "64-400 MIĘDZYCHÓD",
        "address_line2": "UL. SZPITALNA 10",
        "nip": "5951340382",
        "display_name": "Międzychód"
    },
    "limamed": {
        "name": 'Przychodnia Zespołu Lekarza Rodzinnego „Limamed"',
        "address_line1": "64-316 Kuślin",
        "address_line2": "Ul. Emilii Sczanieckiej 6",
        "nip": "7881731812",
        "display_name": "Limamed"
    }
}

MONTHS = [
    "Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec",
    "Lipiec", "Sierpień", "Wrzesień", "Październik", "Listopad", "Grudzień"
]

def create_invoice_excel(invoice_no, date, month, year, hours, rate, total, total_words, buyer_data):
    """Load .xlsx template and replace only data values"""
    import os
    
    # Получаем путь к папке с приложением
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Выбираем шаблон в зависимости от клиники
    if buyer_data.get("display_name") == "Międzychód":
        template_path = os.path.join(base_dir, "shablon", "FakturaSPZOZ Międzychód 22^25.xlsx")
    else:
        template_path = os.path.join(base_dir, "shablon", "Faktura Limamed 23^2025.xlsx")
    
    # Загружаем .xlsx шаблон напрямую через openpyxl (сохраняет все форматирование)
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Меняем ТОЛЬКО данные в нужных ячейках (сохраняем форматирование из шаблона)
    # D9: номер фактуры
    ws['D9'].value = invoice_no
    
    # D16, D18: даты
    ws['D16'].value = date.strftime("%d.%m.%Y")
    ws['D18'].value = date.strftime("%d.%m.%Y")
    
    # C26: описание услуги
    month_name_lower = month.lower()
    ws['C26'].value = f"usługi medyczne lekarza POZ w miesiącu {month_name_lower} {year}"
    
    # E26, F26, G26: часы, ставка, сумма
    ws['E26'].value = int(hours) if hours == int(hours) else hours
    ws['F26'].value = int(rate) if rate == int(rate) else rate
    ws['G26'].value = int(total) if total == int(total) else total
    
    # G27, D28: итого (всегда заменяем на правильную сумму)
    ws['G27'].value = int(total) if total == int(total) else total
    ws['D28'].value = int(total) if total == int(total) else total
    
    # C31: сумма прописью
    ws['C31'].value = total_words
    
    return wb

def main():
    st.title("📄 Generator Faktur")
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Определяем выбранную клинику
    if 'selected_clinic' not in st.session_state:
        st.session_state.selected_clinic = 'miedzychod'
    
    # Выбор клиники - большие кнопки
    st.markdown("### 🏥 Wybierz klinikę")
    
    col1, col2 = st.columns(2)
    
    with col1:
        clinic1_selected = st.session_state.selected_clinic == 'miedzychod'
        btn_type_1 = "primary" if clinic1_selected else "secondary"
        if st.button(f"🏥\n\n{CLINICS['miedzychod']['display_name']}", key="clinic1", use_container_width=True, type=btn_type_1):
            st.session_state.selected_clinic = 'miedzychod'
            st.rerun()
    
    with col2:
        clinic2_selected = st.session_state.selected_clinic == 'limamed'
        btn_type_2 = "primary" if clinic2_selected else "secondary"
        if st.button(f"🏥\n\n{CLINICS['limamed']['display_name']}", key="clinic2", use_container_width=True, type=btn_type_2):
            st.session_state.selected_clinic = 'limamed'
            st.rerun()
    
    selected_clinic_data = CLINICS[st.session_state.selected_clinic]
    st.success(f"✅ Wybrano: **{selected_clinic_data['display_name']}**")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Секция: Данные фактуры
    st.markdown("### 📝 Dane faktury")
    
    # Одна колонка для мобильных
    month = st.selectbox("📅 Miesiąc", MONTHS, index=0, help="Wybierz miesiąc za który wystawiasz fakturę")
    
    invoice_no = st.text_input("🔢 Nr faktury", value="10/2025", help="Format: XX/YYYY (np. 10/2025)")
    
    # Автоматически определяем год
    current_year = datetime.now().year
    st.info(f"📆 **Rok:** {current_year} (ustawiany automatycznie)")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Секция: Usługi
    st.markdown("### 💼 Usługi")
    
    # Используем session_state для хранения значений
    if 'hours_str' not in st.session_state:
        st.session_state.hours_str = ""
    if 'rate' not in st.session_state:
        st.session_state.rate = 170.0
    
    # Поле для часов - text_input чтобы можно было оставить пустым
    hours_input = st.text_input(
        "⏰ Ilość godzin", 
        value=st.session_state.hours_str,
        help="Wprowadź liczbę przepracowanych godzin (np. 111 или 111.5)",
        key="hours_input",
        placeholder="np. 111"
    )
    st.session_state.hours_str = hours_input
    
    # Парсим значение часов
    try:
        hours = float(hours_input.replace(',', '.')) if hours_input.strip() else 0.0
    except ValueError:
        hours = 0.0
        if hours_input.strip():
            st.warning("⚠️ Wprowadź prawidłową liczbę godzin")
    
    rate = st.number_input(
        "💰 Stawka (zł)", 
        min_value=0.0, 
        value=st.session_state.rate,
        step=1.0, 
        format="%.2f",
        help="Stawka za godzinę w złotych",
        key="rate_input"
    )
    st.session_state.rate = rate
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Автоматически используем сегодняшнюю дату
    date = datetime.now().date()
    year = current_year
    
    # Calculate total
    total = hours * rate
    
    # Секция: Podsumowanie
    st.markdown("### 💵 Podsumowanie")
    
    # Display calculation - крупные метрики
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("⏰ Godziny", f"{hours:.1f}")
    with col2:
        st.metric("💰 Stawka", f"{rate:.2f} zł")
    with col3:
        st.metric("💵 RAZEM", f"{total:.2f} zł", delta=None)
    
    # Convert to Polish words
    if total > 0:
        try:
            total_words = num2words(total, lang='pl', to='currency', currency='PLN')
            st.success(f"📝 **Słownie:** {total_words}")
        except Exception as e:
            total_words = f"{total:.2f} zł"
            st.warning(f"Nie udało się przekonwertować na słowa: {e}")
    else:
        total_words = "zero złotych zero groszy"
        st.info("💡 Wprowadź ilość godzin i stawkę, aby zobaczyć podsumowanie")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Валидация перед генерацией
    validation_errors = []
    if not hours_input.strip() or hours <= 0:
        validation_errors.append("⚠️ Wprowadź ilość godzin (większą od 0)")
    if rate <= 0:
        validation_errors.append("⚠️ Wprowadź stawkę (większą od 0)")
    if not invoice_no or invoice_no.strip() == "":
        validation_errors.append("⚠️ Wprowadź numer faktury")
    
    # Показываем ошибки валидации с анимацией (только если есть данные для проверки)
    if hours_input.strip() or rate > 0:
        if validation_errors:
            for error in validation_errors:
                st.warning(error)
    
    # Generate Excel button - большая кнопка с анимацией
    generate_disabled = total <= 0 or len(validation_errors) > 0
    
    if st.button("📥 Generuj fakturę Excel", type="primary", use_container_width=True, disabled=generate_disabled):
        with st.spinner("⏳ Generowanie faktury..."):
            try:
                wb = create_invoice_excel(invoice_no, date, month, year, hours, rate, total, total_words, selected_clinic_data)
                
                # Save to bytes
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                # Download button
                filename = f"Faktura_{invoice_no.replace('/', '_')}.xlsx"
                st.download_button(
                    label="⬇️ Pobierz plik Excel",
                    data=buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
                
                st.balloons()
                st.success(f"✅ Faktura wygenerowana pomyślnie: **{filename}**")
            except Exception as e:
                st.error(f"❌ Błąd podczas generowania faktury: {e}")
                st.exception(e)

if __name__ == "__main__":
    main()

